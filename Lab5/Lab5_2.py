import openpyxl
import matplotlib.pyplot as plt

table = []
c = [i for i in range(11)]
b = [c[i] ** 3 for i in range(11)]
a = [3 * b[i] ** 2 for i in range(11)]

for i in range(11):
    table.append([a[i], b[i], c[i]])

wb = openpyxl.Workbook()

ws = wb.active
ws.title = "Lab5 task 2"

ws['A1'] = "a(b)"
ws['B1'] = "b(c)"
ws['C1'] = "c(a)"
ws['A2'] = "y = x'"
ws['B2'] = "y = x^3"
ws['C2'] = "y = x"


for row in table:
    ws.append(row)

plt.subplots(figsize=(16, 12))
plt.subplot(131)
plt.title("$y = x ^ 3$")
plt.plot(b, a)
plt.subplot(132)
plt.title("$y = x $")
plt.plot(b, b)
plt.subplot(133)
plt.title("$y = \dot x; (y = 3 x ^ 2)$")
plt.plot(a, c)
plt.show()

wb.save('Lab5_2.xlsx')
